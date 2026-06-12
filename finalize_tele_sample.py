import csv
import re
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT = Path("tele_soumu_sample_100_final_plus.csv")
OUTPUT_CSV = Path("tele_soumu_sample_100_final.csv")
OUTPUT_XLSX = Path("tele_soumu_sample_100_final.xlsx")


def normalize_site(url: str) -> str:
    if not url:
        return ""
    parsed = urlparse(url)
    if not parsed.scheme or not parsed.netloc:
        return url
    if parsed.netloc == "www.shinkin.co.jp" and parsed.path.startswith("/chibaskb/"):
        return f"{parsed.scheme}://{parsed.netloc}/chibaskb/"
    return f"{parsed.scheme}://{parsed.netloc}/"


def normalize_fiscal(value: str) -> str:
    if not value or value == "対象外":
        return value
    value = value.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
    match = re.search(r"(\d{1,2})月", value)
    return f"{match.group(1)}月" if match else value


def good_phone(value: str) -> bool:
    if not value:
        return False
    if value in {"03-1234-5678"}:
        return False
    if "-" not in value:
        return False
    if value.startswith("000") or value.startswith("0110-"):
        return False
    return bool(re.fullmatch(r"0\d{1,4}-\d{1,4}-\d{3,4}|0\d{2,4}-\d{3,4}", value))


def main() -> None:
    df = pd.read_csv(INPUT, dtype=str).fillna("")

    for idx, row in df.iterrows():
        df.at[idx, "website_url"] = normalize_site(row.get("website_url", ""))
        df.at[idx, "fiscal_month"] = normalize_fiscal(row.get("fiscal_month", ""))

        if not good_phone(row.get("phone_number", "")):
            df.at[idx, "phone_number"] = ""
            df.at[idx, "phone_source_url"] = ""

        if not good_phone(row.get("fax_number", "")) or row.get("fax_number") == df.at[idx, "phone_number"]:
            df.at[idx, "fax_number"] = ""
            df.at[idx, "fax_source_url"] = ""

        email = row.get("email_address", "")
        if email and email != "info@chizakiroad.co.jp":
            df.at[idx, "email_address"] = ""
            df.at[idx, "email_source_url"] = ""

    front = [
        "company_name",
        "entity_type",
        "corporate_number",
        "detailed_address",
        "phone_number",
        "website_url",
        "fiscal_month",
        "fax_number",
        "email_address",
        "contact_form_url",
        "industry_business",
        "capital",
        "employee_count",
        "office_location",
        "purpose",
        "license_date",
        "station_count",
        "detail_url",
        "nta_source_url",
        "website_source_url",
        "phone_source_url",
        "fax_source_url",
        "email_source_url",
        "contact_form_source_url",
        "capital_source_url",
        "employee_count_source_url",
        "industry_business_source_url",
        "fiscal_month_source_url",
        "irbank_mynumber_url",
        "irbank_company_url",
        "enrichment_status",
    ]
    cols = [c for c in front if c in df.columns] + [c for c in df.columns if c not in front]
    df = df[cols]
    df.to_csv(OUTPUT_CSV, encoding="utf-8-sig", index=False, quoting=csv.QUOTE_MINIMAL)
    df.to_excel(OUTPUT_XLSX, index=False)

    wb = load_workbook(OUTPUT_XLSX)
    ws = wb.active
    ws.title = "sample_100_final"
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col[:101])
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 60)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(OUTPUT_XLSX)

    for col in [
        "detailed_address",
        "website_url",
        "phone_number",
        "fax_number",
        "email_address",
        "contact_form_url",
        "capital",
        "employee_count",
        "fiscal_month",
        "industry_business",
    ]:
        filled = ((df[col] != "") & (df[col] != "対象外")).sum()
        target = (df[col] == "対象外").sum()
        print(col, filled, "target_gai", target)
    print(f"Wrote {OUTPUT_CSV} and {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
