import csv
import re
import time
from pathlib import Path

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT = Path("tele_soumu_sample_100_enriched_contacts.csv")
OUTPUT_CSV = Path("tele_soumu_sample_100_deep.csv")
OUTPUT_XLSX = Path("tele_soumu_sample_100_deep.xlsx")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0 Safari/537.36"
    ),
    "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
}


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def fetch(url: str) -> str:
    try:
        res = requests.get(url, headers=HEADERS, timeout=15)
        res.raise_for_status()
        res.encoding = res.apparent_encoding or res.encoding
        return res.text
    except Exception:
        return ""


def text_from_html(html: str) -> str:
    if not html:
        return ""
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return clean(soup.get_text(" ", strip=True))


def extract_between(text: str, start: str, end_markers: list[str]) -> str:
    i = text.find(start)
    if i == -1:
        return ""
    i += len(start)
    end_positions = [text.find(marker, i) for marker in end_markers]
    end_positions = [pos for pos in end_positions if pos != -1]
    j = min(end_positions) if end_positions else min(len(text), i + 120)
    return clean(text[i:j])


def normalize_phone(value: str) -> str:
    value = value.replace("（", "(").replace("）", ")")
    match = re.search(r"0\d{1,4}[-(]\d{1,4}[-)]\d{3,4}|0\d{1,4}\)\d{1,4}-\d{3,4}", value)
    if not match:
        return ""
    return match.group(0).replace("(", "-").replace(")", "-").strip("-")


def parse_irbank(row: pd.Series) -> dict[str, str]:
    num = str(row.get("corporate_number", "")).strip()
    if not re.fullmatch(r"\d{13}", num):
        return {}
    result: dict[str, str] = {}

    mynumber_url = f"https://irbank.net/mynumber/{num}"
    my_html = fetch(mynumber_url)
    my_text = text_from_html(my_html)
    if not my_text:
        return {}
    result["irbank_mynumber_url"] = mynumber_url

    # Non-listed pages often expose workplace URL, business, and scale.
    url_match = re.search(r"職場情報\s+【URL】\s*(https?://\S+)", my_text)
    if url_match:
        result["website_url"] = url_match.group(1).rstrip("。)")
    business = extract_between(my_text, "【事業】", ["【規模】", "厚生労働省", "【資格】", "政府調達"])
    if business:
        result["industry_business"] = business

    # Listed companies have an E-code page with phone/fiscal/capital detail.
    e_match = re.search(r"関連情報\s+(E\d{5})", my_text)
    if e_match:
        ecode = e_match.group(1)
        e_url = f"https://irbank.net/{ecode}"
        e_text = text_from_html(fetch(e_url))
        if e_text:
            result["irbank_company_url"] = e_url
            phone_raw = extract_between(e_text, "電話番号", ["決算日", "株主総会", "資本金"])
            phone = normalize_phone(phone_raw)
            if phone:
                result["phone_number"] = phone
            fiscal = extract_between(e_text, "決算日", ["株主総会", "資本金", "親会社"])
            if fiscal:
                result["fiscal_month"] = fiscal
            capital = extract_between(
                e_text,
                "資本金",
                ["親会社", "事業の概況", "セグメント", "投資額", "株主情報", "大量保有", "IRBANK"],
            )
            if capital:
                result["capital"] = capital
            industry = extract_between(e_text, "業種", ["セグメント", "投資額", "10年前比較", "社員の状況"])
            overview = extract_between(e_text, "会社情報", ["事業の内容", "企業の概要"])
            parts = [p for p in [industry, overview] if p]
            if parts:
                result["industry_business"] = " / ".join(parts)
            employee = extract_between(e_text, "従業員数", ["平均年齢", "平均勤続年数", "平均年間給与", "役員"])
            if employee:
                result["employee_count"] = employee

    return result


def safe_set(
    df: pd.DataFrame,
    idx: int,
    field: str,
    value: str,
    source_field: str,
    source: str,
    override: bool = False,
) -> None:
    if not value:
        return
    current = str(df.at[idx, field])
    if current and current != "対象外" and not override:
        return
    df.at[idx, field] = value
    if source_field:
        df.at[idx, source_field] = source


def reorder_and_write(df: pd.DataFrame) -> None:
    front = [
        "company_name",
        "entity_type",
        "corporate_number",
        "detailed_address",
        "phone_number",
        "website_url",
        "fiscal_month",
        "fax_number",
        "email_address",
        "contact_form_url",
        "industry_business",
        "capital",
        "employee_count",
        "office_location",
        "purpose",
        "license_date",
        "station_count",
        "detail_url",
        "nta_source_url",
        "website_source_url",
        "phone_source_url",
        "fax_source_url",
        "email_source_url",
        "contact_form_source_url",
        "capital_source_url",
        "employee_count_source_url",
        "industry_business_source_url",
        "fiscal_month_source_url",
        "enrichment_status",
    ]
    cols = [c for c in front if c in df.columns] + [c for c in df.columns if c not in front]
    df = df[cols]
    df.to_csv(OUTPUT_CSV, encoding="utf-8-sig", index=False, quoting=csv.QUOTE_MINIMAL)
    df.to_excel(OUTPUT_XLSX, index=False)

    wb = load_workbook(OUTPUT_XLSX)
    ws = wb.active
    ws.title = "sample_100_deep"
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col[:101])
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 60)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(OUTPUT_XLSX)


def main() -> None:
    df = pd.read_csv(INPUT, dtype=str).fillna("")
    for extra in ["irbank_mynumber_url", "irbank_company_url"]:
        if extra not in df.columns:
            df[extra] = ""

    for idx, row in df.iterrows():
        data = parse_irbank(row)
        source = data.get("irbank_company_url") or data.get("irbank_mynumber_url", "")
        listed_source = bool(data.get("irbank_company_url"))
        safe_set(df, idx, "website_url", data.get("website_url", ""), "website_source_url", source)
        safe_set(df, idx, "phone_number", data.get("phone_number", ""), "phone_source_url", source, override=listed_source)
        safe_set(df, idx, "fiscal_month", data.get("fiscal_month", ""), "fiscal_month_source_url", source, override=listed_source)
        safe_set(df, idx, "capital", data.get("capital", ""), "capital_source_url", source, override=listed_source)
        safe_set(df, idx, "employee_count", data.get("employee_count", ""), "employee_count_source_url", source, override=listed_source)
        safe_set(df, idx, "industry_business", data.get("industry_business", ""), "industry_business_source_url", source, override=listed_source)
        df.at[idx, "irbank_mynumber_url"] = data.get("irbank_mynumber_url", "")
        df.at[idx, "irbank_company_url"] = data.get("irbank_company_url", "")
        print(idx + 1, row["company_name"], data)
        time.sleep(0.4)

    reorder_and_write(df)
    print(f"Wrote {OUTPUT_CSV} and {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
