from __future__ import annotations

import argparse
import csv
import re
import time
import unicodedata
import urllib.parse
from collections import deque
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Iterable

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import urllib3

    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except Exception:
    urllib3 = None

try:
    from PyPDF2 import PdfReader
except Exception:
    PdfReader = None


SOUmu_BASE = "https://www.tele.soumu.go.jp/musen/"
SOUmu_SEARCH = urllib.parse.urljoin(SOUmu_BASE, "SearchServlet")
NTA_URL = "https://www.houjin-bangou.nta.go.jp/henkorireki-johoto.html?selHouzinNo={}"
GBIZ_URL = "https://info.gbiz.go.jp/hojin/ichiran?hojinBango={}"
IRBANK_MYNUMBER_URL = "https://irbank.net/mynumber/{}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0 Safari/537.36"
    ),
    "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
}

SOUmu_Q3_897_5_PARAMS = {
    "SC": "1",
    "pageID": "3",
    "SelectID": "5",
    "CONFIRM": "1",
    "OW": "ML 1",
    "IT": "A",
    "FF": "897.5",
    "TF": "897.5",
    "HZ": "3",
    "DC": "100",
}

EXTRA_CORPORATE_NUMBERS = {
    "株式会社Ｌ＆Ｆアセットファイナンス": "9010001060224",
    "三井住友トラスト・パナソニックファイナンス株式会社": "1010001146146",
    "株式会社三越伊勢丹ホールディングス": "3011101060499",
    "東急不動産株式会社": "7011001016580",
}

# Hints verified while building the 100-row sample. They are used only when public
# databases do not expose a website URL.
KNOWN_OFFICIAL_URLS = {
    "水戸市": "https://www.city.mito.lg.jp/",
    "鉾田市": "https://www.city.hokota.lg.jp/",
    "武州瓦斯株式会社": "https://www.bushugas.co.jp/",
    "蕨市": "https://www.city.warabi.saitama.jp/",
    "株式会社ＺＯＺＯ": "https://corp.zozo.com/",
    "松戸市": "https://www.city.matsudo.chiba.jp/",
    "ジャパンリアルエステイトアセットマネジメント株式会社": "https://www.j-re.co.jp/",
    "ハイウエイ開発株式会社": "https://highway-k.subaru-kougyou.jp/",
    "メルコビルエンジニアリング株式会社": "https://www.resco.co.jp/",
    "リンク情報システム株式会社": "https://www.lis.co.jp/",
    "外務省": "https://www.mofa.go.jp/mofaj/",
    "学校法人日本大学": "https://www.nihon-u.ac.jp/",
    "株式会社全銀電子債権ネットワーク": "https://www.densai.net/",
    "三菱商事都市開発株式会社": "https://www.mcud.co.jp/",
    "田中電気株式会社": "https://www.tanaka-denki.co.jp/",
    "復興庁": "https://www.reconstruction.go.jp/",
    "ＮＴＴアーバンバリューサポート株式会社": "https://www.ntt-uvs.com/",
    "楽天カード株式会社": "https://www.rakuten-card.co.jp/",
    "楽天ペイメント株式会社": "https://payment.rakuten.co.jp/",
    "株式会社ＥＮＥＯＳモビリニア": "https://www.eneos-mobilineer.com/",
    "株式会社長谷工ライブネット": "https://www.haseko-hln.co.jp/",
    "地崎道路株式会社": "https://www.chizakiroad.co.jp/",
    "ＳＯＭＰＯホールディングス株式会社": "https://www.sompo-hd.com/",
    "ベル・データ株式会社": "https://www.belldata.com/",
    "一般財団法人移動無線センター": "https://www.mrc.or.jp/",
    "株式会社三越伊勢丹ホールディングス": "https://www.imhds.co.jp/",
    "公益財団法人東京都公園協会": "https://www.tokyo-park.or.jp/",
    "日清食品ホールディングス株式会社": "https://www.nissin.com/jp/",
    "日本ロレアル株式会社": "https://www.loreal.com/ja-jp/japan/",
    "公益財団法人東京動物園協会": "https://www.tokyo-zoo.net/",
    "台東区": "https://www.city.taito.lg.jp/",
    "社会福祉法人同愛記念病院財団": "https://www.doai.jp/",
    "大田区": "https://www.city.ota.tokyo.jp/",
    "渋谷区": "https://www.city.shibuya.tokyo.jp/",
    "全国農業協同組合連合会": "https://www.zennoh.or.jp/",
    "中野区": "https://www.city.tokyo-nakano.lg.jp/",
    "北区": "https://www.city.kita.tokyo.jp/",
}

CONTACT_KEYWORDS = [
    "お問い合わせ",
    "お問合せ",
    "問合せ",
    "問い合わせ",
    "ご相談",
    "連絡先",
    "窓口",
    "contact",
    "inquiry",
    "otoiawase",
    "toiawase",
    "form",
]

PROFILE_KEYWORDS = [
    "会社概要",
    "企業情報",
    "会社情報",
    "法人概要",
    "団体概要",
    "概要",
    "プロフィール",
    "アクセス",
    "事業所",
    "拠点",
    "IR",
    "投資家",
    "決算",
    "採用",
    "about",
    "company",
    "corporate",
    "profile",
    "outline",
    "access",
    "office",
    "ir",
    "recruit",
]

PDF_KEYWORDS = [
    "会社案内",
    "会社概要",
    "企業情報",
    "決算",
    "統合報告",
    "有価証券報告書",
    "アニュアルレポート",
    "profile",
    "company",
    "outline",
    "report",
    "annual",
    "ir",
]

COMMON_PATHS = [
    "/",
    "/contact/",
    "/inquiry/",
    "/otoiawase/",
    "/toiawase/",
    "/company/",
    "/company/outline/",
    "/company/profile/",
    "/corporate/",
    "/corporate/profile/",
    "/about/",
    "/profile/",
    "/outline/",
    "/access/",
    "/ir/",
    "/recruit/",
]

EXCLUDED_SEARCH_DOMAINS = [
    "houjin-bangou.nta.go.jp",
    "info.gbiz.go.jp",
    "irbank.net",
    "buffett-code.com",
    "baseconnect.in",
    "salesnow.jp",
    "catr.jp",
    "initial.inc",
    "job.mynavi.jp",
    "mynavi.jp",
    "rikunabi",
    "doda.jp",
    "en-gage.net",
    "wantedly.com",
    "green-japan.com",
    "openwork.jp",
    "vorkers.com",
    "indeed.com",
    "prtimes.jp",
    "facebook.com",
    "twitter.com",
    "x.com",
    "linkedin.com",
    "instagram.com",
    "wikipedia.org",
    "mapion.co.jp",
    "navitime.co.jp",
    "map.yahoo.co.jp",
]

SOURCE_COLUMNS = {
    "detailed_address": "address_source_url",
    "phone_number": "phone_source_url",
    "website_url": "website_source_url",
    "fiscal_month": "fiscal_month_source_url",
    "fax_number": "fax_source_url",
    "email_address": "email_source_url",
    "contact_form_url": "contact_form_source_url",
    "industry_business": "industry_business_source_url",
    "capital": "capital_source_url",
    "employee_count": "employee_count_source_url",
}

CONFIDENCE_COLUMNS = {
    "detailed_address": "address_confidence",
    "phone_number": "phone_confidence",
    "website_url": "website_confidence",
    "fiscal_month": "fiscal_month_confidence",
    "fax_number": "fax_confidence",
    "email_address": "email_confidence",
    "contact_form_url": "contact_form_confidence",
    "industry_business": "industry_business_confidence",
    "capital": "capital_confidence",
    "employee_count": "employee_count_confidence",
}

FINAL_FRONT_COLUMNS = [
    "company_name",
    "entity_type",
    "corporate_number",
    "detailed_address",
    "phone_number",
    "website_url",
    "fiscal_month",
    "fax_number",
    "email_address",
    "contact_form_url",
    "industry_business",
    "capital",
    "employee_count",
    "office_location",
    "purpose",
    "license_date",
    "station_count",
    "detail_url",
    "address_source_url",
    "website_source_url",
    "phone_source_url",
    "fax_source_url",
    "email_source_url",
    "contact_form_source_url",
    "capital_source_url",
    "employee_count_source_url",
    "industry_business_source_url",
    "fiscal_month_source_url",
    "address_confidence",
    "website_confidence",
    "phone_confidence",
    "fax_confidence",
    "email_confidence",
    "contact_form_confidence",
    "capital_confidence",
    "employee_count_confidence",
    "industry_business_confidence",
    "fiscal_month_confidence",
    "nta_source_url",
    "gbiz_source_url",
    "irbank_mynumber_url",
    "irbank_company_url",
    "enrichment_status",
    "manual_check_reason",
]


@dataclass
class FetchResult:
    requested_url: str
    final_url: str
    html: str
    status_code: int
    error: str = ""


class Fetcher:
    def __init__(self, sleep_seconds: float, timeout: int = 18, retries: int = 2) -> None:
        self.sleep_seconds = sleep_seconds
        self.timeout = timeout
        self.retries = retries
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self._last_request_at = 0.0

    def fetch(self, url: str, *, allow_insecure_retry: bool = True) -> FetchResult:
        url = ensure_url(url)
        if not url:
            return FetchResult(url, url, "", 0, "empty url")

        last_error = ""
        for attempt in range(self.retries + 1):
            self._throttle()
            try:
                response = self.session.get(
                    url,
                    timeout=self.timeout,
                    allow_redirects=True,
                    verify=True,
                )
            except requests.exceptions.SSLError as exc:
                if not allow_insecure_retry:
                    last_error = str(exc)
                    continue
                try:
                    self._throttle()
                    response = self.session.get(
                        url,
                        timeout=self.timeout,
                        allow_redirects=True,
                        verify=False,
                    )
                except Exception as inner_exc:
                    last_error = str(inner_exc)
                    time.sleep(0.6 * (attempt + 1))
                    continue
            except Exception as exc:
                last_error = str(exc)
                time.sleep(0.6 * (attempt + 1))
                continue

            status = response.status_code
            if status >= 500 and attempt < self.retries:
                time.sleep(0.8 * (attempt + 1))
                continue
            if status >= 400:
                return FetchResult(url, response.url, "", status, f"HTTP {status}")
            response.encoding = response.apparent_encoding or response.encoding
            return FetchResult(url, response.url, response.text, status, "")

        return FetchResult(url, url, "", 0, last_error)

    def fetch_bytes(self, url: str, *, max_bytes: int = 15_000_000) -> tuple[bytes, str, str]:
        url = ensure_url(url)
        if not url:
            return b"", url, "empty url"
        last_error = ""
        for attempt in range(self.retries + 1):
            self._throttle()
            try:
                response = self.session.get(
                    url,
                    timeout=self.timeout,
                    allow_redirects=True,
                    verify=True,
                )
            except requests.exceptions.SSLError:
                try:
                    self._throttle()
                    response = self.session.get(
                        url,
                        timeout=self.timeout,
                        allow_redirects=True,
                        verify=False,
                    )
                except Exception as exc:
                    last_error = str(exc)
                    time.sleep(0.6 * (attempt + 1))
                    continue
            except Exception as exc:
                last_error = str(exc)
                time.sleep(0.6 * (attempt + 1))
                continue
            if response.status_code >= 400:
                return b"", response.url, f"HTTP {response.status_code}"
            content = response.content
            if len(content) > max_bytes:
                return b"", response.url, f"too large: {len(content)} bytes"
            return content, response.url, ""
        return b"", url, last_error

    def _throttle(self) -> None:
        elapsed = time.time() - self._last_request_at
        if elapsed < self.sleep_seconds:
            time.sleep(self.sleep_seconds - elapsed)
        self._last_request_at = time.time()


def normalize_text(text: str) -> str:
    if text is None:
        return ""
    text = unicodedata.normalize("NFKC", str(text))
    text = text.replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean(text: str) -> str:
    return normalize_text(text)


def ensure_url(url: str) -> str:
    url = clean(url)
    if not url:
        return ""
    if url.startswith("//"):
        return "https:" + url
    if not re.match(r"https?://", url):
        return "https://" + url
    return url


def root_url(url: str) -> str:
    url = ensure_url(url)
    parsed = urllib.parse.urlparse(url)
    if not parsed.scheme or not parsed.netloc:
        return url
    if parsed.netloc == "www.shinkin.co.jp" and parsed.path.startswith("/chibaskb/"):
        return f"{parsed.scheme}://{parsed.netloc}/chibaskb/"
    return f"{parsed.scheme}://{parsed.netloc}/"


def is_valid_http_url(url: str) -> bool:
    url = ensure_url(url)
    parsed = urllib.parse.urlparse(url)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def known_official_url(company_name: str) -> str:
    normalized_name = clean(company_name)
    for key, value in KNOWN_OFFICIAL_URLS.items():
        if clean(key) == normalized_name:
            return value
    return ""


def strip_source_note(text: str) -> str:
    text = re.sub(
        r"\s*[（(]\s*[^（）()]*?(法人番号公表サイト|職場情報総合サイト|EDINET|GEPS)[^（）()]*?\s*[）)]",
        "",
        text,
    )
    return clean(text)


def first_nonempty(*values: str) -> str:
    for value in values:
        value = clean(value)
        if value:
            return value
    return ""


def build_soumu_url(params: dict[str, str]) -> str:
    query = urllib.parse.urlencode(params).replace("+", "%20")
    return f"{SOUmu_SEARCH}?{query}#result"


def table_rows(soup: BeautifulSoup) -> list[list[str]]:
    rows: list[list[str]] = []
    table = soup.find("table")
    if not table:
        return rows
    for tr in table.find_all("tr"):
        cells = [clean(cell.get_text(" ", strip=True)) for cell in tr.find_all(["th", "td"])]
        if cells:
            rows.append(cells)
    return rows


def parse_name_and_corporate_number(raw_name: str) -> tuple[str, str]:
    match = re.search(r"^(.*?)\s+.*?(\d{13})$", clean(raw_name))
    if not match:
        return clean(raw_name), ""
    return match.group(1).strip(), match.group(2)


def parse_soumu_detail(fetcher: Fetcher, url: str) -> dict[str, str]:
    result = fetcher.fetch(url)
    soup = BeautifulSoup(result.html, "html.parser")
    rows = table_rows(soup)

    def cell(row_index: int, col_index: int = 1) -> str:
        if row_index >= len(rows):
            return ""
        row = rows[row_index]
        return row[col_index] if len(row) > col_index else ""

    return {
        "detail_licensee_name": cell(0),
        "detail_address": cell(1),
        "station_type": cell(2),
        "license_number": cell(3, 3),
        "valid_period": cell(4, 3),
        "station_count": cell(5),
        "office_location_detail": cell(7),
    }


def scrape_soumu_q3_897_5(fetcher: Fetcher, limit: int) -> list[dict[str, str]]:
    result = fetcher.fetch(build_soumu_url(SOUmu_Q3_897_5_PARAMS))
    if not result.html:
        raise RuntimeError(f"Could not fetch Soumu search result: {result.error}")
    soup = BeautifulSoup(result.html, "html.parser")
    records: list[dict[str, str]] = []
    for tr in soup.select("tr.m-table-sort__row"):
        cells = [clean(cell.get_text(" ", strip=True)) for cell in tr.find_all("td")]
        link = tr.select_one('a[href*="pageID=4"]')
        if len(cells) < 4 or not link:
            continue
        detail_url = urllib.parse.urljoin(SOUmu_BASE, link["href"])
        company_name, corporate_number = parse_name_and_corporate_number(cells[0])
        records.append(
            {
                "source_query": "q3_kanto_land_mobile_blanket_897_5mhz",
                "company_name": company_name,
                "detail_licensee_name": cells[0],
                "corporate_number": corporate_number,
                "office_location": cells[1],
                "purpose": cells[2],
                "license_date": cells[3],
                "detail_url": detail_url,
                **parse_soumu_detail(fetcher, detail_url),
            }
        )
        print(f"scraped {len(records):03d}: {company_name}")
        if len(records) >= limit:
            break
    return records


def parse_nta_profile(fetcher: Fetcher, corporate_number: str) -> dict[str, str]:
    if not re.fullmatch(r"\d{13}", corporate_number or ""):
        return {}
    url = NTA_URL.format(corporate_number)
    result = fetcher.fetch(url)
    if not result.html:
        return {"nta_source_url": url, "nta_error": result.error}
    soup = BeautifulSoup(result.html, "html.parser")
    profile: dict[str, str] = {}
    for dt in soup.find_all("dt"):
        dd = dt.find_next_sibling("dd")
        if not dd:
            continue
        key = clean(dt.get_text(" ", strip=True))
        value = clean(dd.get_text(" ", strip=True))
        if key and value:
            profile[key] = value
    return {
        "nta_name": profile.get("商号又は名称", ""),
        "detailed_address": profile.get("本店又は主たる事務所の所在地", ""),
        "nta_last_updated": profile.get("最終更新年月日", ""),
        "nta_source_url": url,
    }


def parse_gbiz_profile(fetcher: Fetcher, corporate_number: str) -> dict[str, str]:
    if not re.fullmatch(r"\d{13}", corporate_number or ""):
        return {}
    url = GBIZ_URL.format(corporate_number)
    result = fetcher.fetch(url)
    if not result.html:
        return {"gbiz_source_url": url, "gbiz_error": result.error}
    soup = BeautifulSoup(result.html, "html.parser")
    data: dict[str, str] = {"gbiz_source_url": url}

    for body in soup.select(".accordion-body"):
        for item in body.select(".row"):
            children = item.find_all(recursive=False)
            if len(children) < 2:
                continue
            key = clean(children[0].get_text(" ", strip=True))
            value = strip_source_note(children[1].get_text(" ", strip=True))
            if key and value and key not in data:
                data[key] = value
        if len(data) > 1:
            break

    text = clean(soup.get_text(" ", strip=True))
    match = re.search(r"当期\s+第[^（]*（自\s+\d{4}年\d{1,2}月\d{1,2}日\s+至\s+\d{4}年(\d{1,2})月\d{1,2}日", text)
    if match:
        data["決算月"] = f"{match.group(1)}月"
    return data


def extract_between(text: str, start: str, end_markers: Iterable[str], max_chars: int = 180) -> str:
    i = text.find(start)
    if i == -1:
        return ""
    i += len(start)
    end_positions = [text.find(marker, i) for marker in end_markers]
    end_positions = [pos for pos in end_positions if pos != -1]
    j = min(end_positions) if end_positions else min(len(text), i + max_chars)
    return clean(text[i:j])


def parse_irbank(fetcher: Fetcher, corporate_number: str) -> dict[str, str]:
    if not re.fullmatch(r"\d{13}", corporate_number or ""):
        return {}
    result: dict[str, str] = {}
    mynumber_url = IRBANK_MYNUMBER_URL.format(corporate_number)
    my_result = fetcher.fetch(mynumber_url)
    my_text = visible_text(my_result.html)
    if not my_text:
        return {}
    result["irbank_mynumber_url"] = mynumber_url

    url_match = re.search(r"職場情報\s+【URL】\s*(https?://\S+)", my_text)
    if url_match:
        result["website_url"] = url_match.group(1).rstrip("。)")
    business = extract_between(my_text, "【事業】", ["【規模】", "厚生労働省", "【資格】", "政府調達"], 400)
    if business:
        result["industry_business"] = business
    scale = extract_between(my_text, "【規模】", ["【資格】", "厚生労働省", "政府調達"], 160)
    if scale and "人" in scale:
        result["employee_count"] = scale

    e_match = re.search(r"関連情報\s+(E\d{5})", my_text)
    if not e_match:
        return result
    ecode = e_match.group(1)
    e_url = f"https://irbank.net/{ecode}"
    e_result = fetcher.fetch(e_url)
    e_text = visible_text(e_result.html)
    if not e_text:
        return result

    result["irbank_company_url"] = e_url
    phone = normalize_phone(extract_between(e_text, "電話番号", ["決算日", "株主総会", "資本金"], 80))
    if phone:
        result["phone_number"] = phone
    fiscal = parse_fiscal_month(extract_between(e_text, "決算日", ["株主総会", "資本金", "親会社"], 80))
    if fiscal:
        result["fiscal_month"] = fiscal
    capital = extract_between(
        e_text,
        "資本金",
        ["親会社", "事業の概況", "セグメント", "投資額", "株主情報", "大量保有", "IRBANK"],
        120,
    )
    if capital:
        result["capital"] = capital
    industry = extract_between(e_text, "業種", ["セグメント", "投資額", "10年前比較", "社員の状況"], 160)
    overview = extract_between(e_text, "会社情報", ["事業の内容", "企業の概要"], 220)
    industry_parts = [part for part in [industry, overview] if part]
    if industry_parts:
        result["industry_business"] = " / ".join(industry_parts)
    employees = extract_between(e_text, "従業員数", ["平均年齢", "平均勤続年数", "平均年間給与", "役員"], 100)
    if employees:
        result["employee_count"] = employees
    return result


def classify_entity(name: str) -> str:
    name = clean(name)
    if re.search(r"(市|区|町|村)$", name):
        return "自治体"
    if name.endswith("省") or name.endswith("庁"):
        return "官公庁"
    if "学校法人" in name:
        return "学校法人"
    if "社会福祉法人" in name:
        return "社会福祉法人"
    if "公益財団法人" in name or "一般財団法人" in name:
        return "財団法人"
    if "生活協同組合" in name or "信用金庫" in name or name.endswith("連合会"):
        return "その他法人"
    return "株式会社等"


def not_applicable(entity_type: str, field: str) -> str:
    if field == "capital" and entity_type in {"自治体", "官公庁", "学校法人", "社会福祉法人", "財団法人", "その他法人"}:
        return "対象外"
    if field == "fiscal_month" and entity_type in {"自治体", "官公庁"}:
        return "対象外"
    return ""


def visible_text(html: str) -> str:
    if not html:
        return ""
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    return clean(soup.get_text(" ", strip=True))


def extract_key_values(soup: BeautifulSoup) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []

    for tr in soup.find_all("tr"):
        ths = tr.find_all("th")
        tds = tr.find_all("td")
        if ths and tds:
            key = clean(" ".join(th.get_text(" ", strip=True) for th in ths))
            value = clean(" ".join(td.get_text(" ", strip=True) for td in tds))
            if key and value:
                pairs.append((key, value))
            continue
        cells = [clean(cell.get_text(" ", strip=True)) for cell in tr.find_all(["td", "th"])]
        if len(cells) >= 2 and is_label_like(cells[0]):
            pairs.append((cells[0], clean(" ".join(cells[1:]))))

    for dl in soup.find_all("dl"):
        dts = dl.find_all("dt")
        for dt in dts:
            dd = dt.find_next_sibling("dd")
            if not dd:
                continue
            key = clean(dt.get_text(" ", strip=True))
            value = clean(dd.get_text(" ", strip=True))
            if key and value:
                pairs.append((key, value))

    for tag in soup.find_all(["p", "li", "div"]):
        text = clean(tag.get_text(" ", strip=True))
        if len(text) > 180:
            continue
        match = re.match(r"^(.{2,18}?)[：:]\s*(.+)$", text)
        if match and is_label_like(match.group(1)):
            pairs.append((clean(match.group(1)), clean(match.group(2))))

    return pairs


def is_label_like(text: str) -> bool:
    labels = [
        "住所",
        "所在地",
        "本店所在地",
        "本社所在地",
        "電話",
        "TEL",
        "FAX",
        "資本金",
        "従業員",
        "社員数",
        "職員数",
        "事業内容",
        "業務内容",
        "営業品目",
        "決算",
    ]
    upper = clean(text).upper()
    return any(label.upper() in upper for label in labels)


def find_pair_value(pairs: list[tuple[str, str]], labels: Iterable[str]) -> str:
    label_list = [clean(label).upper() for label in labels]
    for key, value in pairs:
        key_norm = clean(key).upper()
        if any(label in key_norm for label in label_list):
            return trim_value(value)
    return ""


def trim_value(value: str, max_len: int = 300) -> str:
    value = clean(value)
    value = re.sub(r"\s+(詳細はこちら|詳しくはこちら|MORE|more)\s*$", "", value)
    if len(value) > max_len:
        value = value[:max_len].rstrip() + "..."
    return value


def normalize_phone(value: str) -> str:
    value = normalize_text(value)
    value = value.replace("ー", "-").replace("−", "-").replace("－", "-").replace("―", "-").replace("‐", "-")
    value = value.replace("(", "-").replace(")", "-")
    match = re.search(r"0\d{1,4}[-]?\d{1,4}[-]?\d{3,4}", value)
    if not match:
        return ""
    raw = match.group(0).strip("-")
    digits = re.sub(r"\D", "", raw)
    if len(digits) not in {10, 11}:
        return ""
    if digits.startswith(("000", "0110")):
        return ""
    if "-" in raw:
        raw = re.sub(r"-+", "-", raw).strip("-")
        if good_phone(raw):
            return raw
    if len(digits) == 10 and digits.startswith(("03", "04", "06")):
        return f"{digits[:2]}-{digits[2:6]}-{digits[6:]}"
    if len(digits) == 10 and digits.startswith("0120"):
        return f"{digits[:4]}-{digits[4:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    if len(digits) == 11 and digits.startswith(("050", "070", "080", "090")):
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 11 and digits.startswith(("0120", "0800")):
        return f"{digits[:4]}-{digits[4:7]}-{digits[7:]}"
    return ""


def good_phone(value: str) -> bool:
    value = clean(value)
    if not value or value == "03-1234-5678":
        return False
    digits = re.sub(r"\D", "", value)
    if len(digits) not in {10, 11}:
        return False
    if digits.startswith(("000", "0110")):
        return False
    return bool(re.fullmatch(r"0\d{1,4}-\d{1,4}-\d{3,4}", value))


def extract_phone_and_fax(text: str, soup: BeautifulSoup) -> tuple[str, str]:
    text = normalize_text(text)
    phone_candidates: list[tuple[int, str]] = []
    fax_candidates: list[tuple[int, str]] = []

    for link in soup.find_all("a", href=True):
        href = clean(link["href"])
        if href.lower().startswith("tel:"):
            phone = normalize_phone(href.split(":", 1)[1])
            if phone:
                phone_candidates.append((100, phone))

    phone_pattern = re.compile(r"0\d{1,4}[\-ー－―‐(]?\d{1,4}[\-ー－―‐)]?\d{3,4}")
    for match in phone_pattern.finditer(text):
        phone = normalize_phone(match.group(0))
        if not phone:
            continue
        context = text[max(0, match.start() - 80) : min(len(text), match.end() + 60)].upper()
        if "FAX" in context or "ファックス" in context:
            fax_candidates.append((score_contact_context(context, is_fax=True), phone))
        else:
            score = score_contact_context(context, is_fax=False)
            if score > 0:
                phone_candidates.append((score, phone))

    phone = best_scored_value(phone_candidates)
    fax = best_scored_value(fax_candidates)
    if phone and fax == phone:
        fax = ""
    return phone, fax


def score_contact_context(context: str, *, is_fax: bool) -> int:
    score = 1
    positive = ["TEL", "電話", "代表", "本社", "連絡先", "お問い合わせ", "問合せ"]
    fax_positive = ["FAX", "ファックス"]
    negative = ["採用", "個人情報", "報道", "IR", "株主", "緊急", "警察"]
    if is_fax:
        for word in fax_positive:
            if word in context:
                score += 8
    else:
        for word in positive:
            if word in context:
                score += 4
        for word in fax_positive:
            if word in context:
                score -= 8
    for word in negative:
        if word in context:
            score -= 2
    return score


def best_scored_value(candidates: list[tuple[int, str]]) -> str:
    if not candidates:
        return ""
    candidates = [(score, value) for score, value in candidates if value and score > 0]
    if not candidates:
        return ""
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def extract_emails(text: str, soup: BeautifulSoup) -> str:
    candidates: list[str] = []
    for link in soup.find_all("a", href=True):
        href = link["href"].strip()
        if href.lower().startswith("mailto:"):
            candidates.append(href.split(":", 1)[1].split("?", 1)[0])

    normalized = normalize_text(text)
    normalized = normalized.replace("＠", "@")
    normalized = re.sub(r"\s*(\[at\]|\(at\)| at |＠)\s*", "@", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"\s*(\[dot\]|\(dot\)| dot )\s*", ".", normalized, flags=re.IGNORECASE)
    candidates.extend(re.findall(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", normalized))

    for email in candidates:
        email = clean(email).strip("。.,;:）)]】")
        if is_good_email(email):
            return email
    return ""


def is_good_email(email: str) -> bool:
    if not email or "@" not in email:
        return False
    lower = email.lower()
    bad_fragments = ["example.", "sample", "dummy", "xxxx", "xxx@", "no-reply", "noreply"]
    if any(fragment in lower for fragment in bad_fragments):
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", email))


def parse_fiscal_month(value: str) -> str:
    value = normalize_text(value)
    if not value:
        return ""
    match = re.search(r"(\d{1,2})月", value)
    if match:
        month = int(match.group(1))
        if 1 <= month <= 12:
            return f"{month}月"
    return ""


def extract_contact_form_url(soup: BeautifulSoup, current_url: str) -> str:
    parsed_current = urllib.parse.urlparse(current_url)
    if soup.find("form") and has_keyword(current_url, CONTACT_KEYWORDS):
        return current_url
    scored: list[tuple[int, str]] = []
    for link in soup.find_all("a", href=True):
        label = clean(link.get_text(" ", strip=True))
        href = clean(link["href"])
        if not href or href.lower().startswith(("tel:", "mailto:", "javascript:")):
            continue
        url = urllib.parse.urljoin(current_url, href).split("#")[0]
        haystack = f"{label} {href}"
        if not has_keyword(haystack, CONTACT_KEYWORDS):
            continue
        parsed = urllib.parse.urlparse(url)
        score = 10
        if parsed.netloc and parsed.netloc != parsed_current.netloc:
            score -= 2
        if "contact" in url.lower() or "inquiry" in url.lower():
            score += 2
        scored.append((score, url))
    if not scored:
        return ""
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored[0][1]


def has_keyword(text: str, keywords: Iterable[str]) -> bool:
    lower = normalize_text(text).lower()
    return any(keyword.lower() in lower for keyword in keywords)


def extract_fields_from_html(html: str, final_url: str) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    text = clean(soup.get_text(" ", strip=True))
    pairs = extract_key_values(soup)
    phone, fax = extract_phone_and_fax(text, soup)

    capital = first_nonempty(
        find_pair_value(pairs, ["資本金", "出資金"]),
        regex_extract(text, r"資本金\s*[:：]?\s*([0-9,億千万百万円円\s]+)", 80),
    )
    employee_count = first_nonempty(
        find_pair_value(pairs, ["従業員数", "従業員", "社員数", "職員数", "従業者数"]),
        regex_extract(text, r"(?:従業員数|従業員|社員数|職員数)\s*[:：]?\s*([0-9,約名人\s]+)", 80),
    )
    fiscal_month = first_nonempty(
        parse_fiscal_month(find_pair_value(pairs, ["決算期", "決算月", "決算日", "事業年度"])),
        parse_fiscal_month(regex_extract(text, r"(?:決算期|決算月|決算日|事業年度)\s*[:：]?\s*([^。]{1,80})", 80)),
    )
    business = first_nonempty(
        find_pair_value(pairs, ["事業内容", "事業概要", "業務内容", "営業品目", "主な事業"]),
        regex_extract(text, r"(?:事業内容|事業概要|業務内容|営業品目)\s*[:：]?\s*([^。]{6,250})", 250),
    )
    address = first_nonempty(
        find_pair_value(pairs, ["本社所在地", "本店所在地", "所在地", "住所"]),
        regex_extract(text, r"(?:本社所在地|本店所在地|所在地|住所)\s*[:：]?\s*([^。]{8,120})", 120),
    )

    return {
        "detailed_address": address,
        "phone_number": phone,
        "fax_number": fax,
        "email_address": extract_emails(text, soup),
        "contact_form_url": extract_contact_form_url(soup, final_url),
        "capital": capital,
        "employee_count": employee_count,
        "fiscal_month": fiscal_month,
        "industry_business": business,
    }


def extract_fields_from_text(text: str) -> dict[str, str]:
    text = normalize_text(text)
    empty_soup = BeautifulSoup("", "html.parser")
    phone, fax = extract_phone_and_fax(text, empty_soup)
    capital = regex_extract(text, r"資本金\s*[:：]?\s*([0-9,億千万百万円円\s]+)", 80)
    employee_count = regex_extract(text, r"(?:従業員数|従業員|社員数|職員数)\s*[:：]?\s*([0-9,約名人\s]+)", 80)
    fiscal_month = parse_fiscal_month(
        regex_extract(text, r"(?:決算期|決算月|決算日|事業年度)\s*[:：]?\s*([^。]{1,80})", 80)
    )
    business = regex_extract(text, r"(?:事業内容|事業概要|業務内容|営業品目)\s*[:：]?\s*([^。]{6,250})", 250)
    address = regex_extract(text, r"(?:本社所在地|本店所在地|所在地|住所)\s*[:：]?\s*([^。]{8,120})", 120)
    return {
        "detailed_address": address,
        "phone_number": phone,
        "fax_number": fax,
        "email_address": extract_emails(text, empty_soup),
        "contact_form_url": "",
        "capital": capital,
        "employee_count": employee_count,
        "fiscal_month": fiscal_month,
        "industry_business": business,
    }


def regex_extract(text: str, pattern: str, max_len: int) -> str:
    match = re.search(pattern, normalize_text(text), flags=re.IGNORECASE)
    if not match:
        return ""
    return trim_value(match.group(1), max_len)


def same_site(base_url: str, target_url: str) -> bool:
    base = urllib.parse.urlparse(base_url)
    target = urllib.parse.urlparse(target_url)
    if not target.netloc:
        return True
    return target.netloc == base.netloc or target.netloc.endswith("." + base.netloc)


def candidate_links(base_url: str, html: str, max_links: int) -> list[str]:
    parsed = urllib.parse.urlparse(base_url)
    origin = f"{parsed.scheme}://{parsed.netloc}"
    candidates: list[tuple[int, str]] = []
    for index, path in enumerate(COMMON_PATHS):
        candidates.append((100 - index, urllib.parse.urljoin(origin, path)))

    soup = BeautifulSoup(html, "html.parser")
    for link in soup.find_all("a", href=True):
        label = clean(link.get_text(" ", strip=True))
        href = clean(link["href"])
        if not href or href.lower().startswith(("mailto:", "tel:", "javascript:")):
            continue
        url = urllib.parse.urljoin(base_url, href).split("#")[0]
        if not same_site(base_url, url):
            continue
        haystack = f"{label} {href}"
        score = 0
        if has_keyword(haystack, CONTACT_KEYWORDS):
            score += 80
        if has_keyword(haystack, PROFILE_KEYWORDS):
            score += 60
        if score:
            candidates.append((score, url))

    seen: set[str] = set()
    ordered: list[str] = []
    for _, url in sorted(candidates, key=lambda item: item[0], reverse=True):
        if url in seen:
            continue
        seen.add(url)
        ordered.append(url)
        if len(ordered) >= max_links:
            break
    return ordered


def candidate_pdf_links(base_url: str, html: str) -> list[tuple[int, str]]:
    links: list[tuple[int, str]] = []
    soup = BeautifulSoup(html, "html.parser")
    for link in soup.find_all("a", href=True):
        label = clean(link.get_text(" ", strip=True))
        href = clean(link["href"])
        if not href or href.lower().startswith(("mailto:", "tel:", "javascript:")):
            continue
        url = urllib.parse.urljoin(base_url, href).split("#")[0]
        if ".pdf" not in urllib.parse.urlparse(url).path.lower():
            continue
        haystack = f"{label} {href}"
        score = 1
        if has_keyword(haystack, PDF_KEYWORDS + PROFILE_KEYWORDS):
            score += 10
        if has_keyword(haystack, ["個人情報", "privacy", "約款", "terms"]):
            score -= 8
        if score > 0:
            links.append((score, url))
    return links


def discover_pdf_urls(fetcher: Fetcher, website_url: str, page_limit: int, pdf_limit: int) -> list[str]:
    if pdf_limit <= 0 or PdfReader is None:
        return []
    start = fetcher.fetch(website_url)
    if not start.html:
        return []

    page_urls = candidate_links(start.final_url, start.html, page_limit)
    if start.final_url not in page_urls:
        page_urls.insert(0, start.final_url)

    scored: list[tuple[int, str]] = []
    visited_pages: set[str] = set()
    for page_url in page_urls[:page_limit]:
        if page_url in visited_pages:
            continue
        visited_pages.add(page_url)
        page = start if page_url == start.final_url else fetcher.fetch(page_url)
        if not page.html:
            continue
        scored.extend(candidate_pdf_links(page.final_url, page.html))

    seen: set[str] = set()
    ordered: list[str] = []
    for _, url in sorted(scored, key=lambda item: item[0], reverse=True):
        if url in seen:
            continue
        seen.add(url)
        ordered.append(url)
        if len(ordered) >= pdf_limit:
            break
    return ordered


def extract_pdf_text(fetcher: Fetcher, pdf_url: str, max_pages: int) -> str:
    if PdfReader is None:
        return ""
    content, _, error = fetcher.fetch_bytes(pdf_url)
    if error or not content:
        return ""
    try:
        reader = PdfReader(BytesIO(content))
        parts: list[str] = []
        for page in reader.pages[:max_pages]:
            try:
                parts.append(page.extract_text() or "")
            except Exception:
                continue
        return clean(" ".join(parts))
    except Exception:
        return ""
    


def crawl_official_site(fetcher: Fetcher, website_url: str, max_pages: int) -> list[tuple[str, dict[str, str]]]:
    website_url = ensure_url(website_url)
    if not website_url:
        return []

    start = fetcher.fetch(website_url)
    if not start.html:
        return []

    queue: deque[str] = deque(candidate_links(start.final_url, start.html, max_pages))
    if start.final_url not in queue:
        queue.appendleft(start.final_url)

    visited: set[str] = set()
    extracted: list[tuple[str, dict[str, str]]] = []
    while queue and len(visited) < max_pages:
        url = queue.popleft()
        if url in visited:
            continue
        visited.add(url)
        result = start if url == start.final_url else fetcher.fetch(url)
        if not result.html:
            continue
        data = extract_fields_from_html(result.html, result.final_url)
        extracted.append((result.final_url, data))
        if len(visited) < max_pages:
            for next_url in candidate_links(result.final_url, result.html, max_pages):
                if next_url not in visited and next_url not in queue:
                    queue.append(next_url)
    return extracted


def company_core_name(name: str) -> str:
    text = normalize_text(name)
    for word in [
        "株式会社",
        "有限会社",
        "合同会社",
        "一般財団法人",
        "公益財団法人",
        "社会福祉法人",
        "学校法人",
        "相互会社",
        "ホールディングス",
        "ＨＤ",
    ]:
        text = text.replace(word, "")
    text = re.sub(r"[\s・･,，.。()（）「」『』\-ー－]", "", text)
    return text


def search_web_for_official_site(fetcher: Fetcher, company_name: str) -> tuple[str, str, str]:
    candidates: list[tuple[int, str, str]] = []
    query = f'"{company_name}" 公式 会社概要'
    encoded = urllib.parse.quote(query)
    search_urls = [
        f"https://duckduckgo.com/html/?q={encoded}",
        f"https://www.bing.com/search?q={encoded}&cc=jp",
    ]
    for search_url in search_urls:
        result = fetcher.fetch(search_url)
        if not result.html:
            continue
        soup = BeautifulSoup(result.html, "html.parser")
        links = soup.select("a.result__a") or soup.select("li.b_algo h2 a")
        for link in links[:8]:
            url = clean(link.get("href", ""))
            title = clean(link.get_text(" ", strip=True))
            if "duckduckgo.com/l/?" in url:
                parsed = urllib.parse.urlparse(url)
                qs = urllib.parse.parse_qs(parsed.query)
                url = qs.get("uddg", [""])[0]
            if not url.startswith("http"):
                continue
            score = score_official_candidate(company_name, url, title)
            if score > 0:
                candidates.append((score, url, title))
    if not candidates:
        return "", "", ""
    candidates.sort(key=lambda item: item[0], reverse=True)
    best_score, best_url, title = candidates[0]
    if best_score < 6:
        return "", "", ""
    return root_url(best_url), best_url, title


def score_official_candidate(company_name: str, url: str, title: str) -> int:
    parsed = urllib.parse.urlparse(url)
    netloc = parsed.netloc.lower().replace("www.", "")
    if any(bad in netloc for bad in EXCLUDED_SEARCH_DOMAINS):
        return -100
    title_norm = normalize_text(title)
    core = company_core_name(company_name)
    score = 0
    if company_name and company_name in title_norm:
        score += 8
    if core and core in normalize_text(f"{title} {url}"):
        score += 6
    if any(word in title_norm for word in ["公式", "会社概要", "企業情報", "ホームページ"]):
        score += 3
    if parsed.netloc.endswith((".co.jp", ".or.jp", ".go.jp", ".lg.jp", ".ac.jp")):
        score += 2
    if any(word in parsed.path.lower() for word in ["company", "corporate", "about", "profile", "outline"]):
        score += 1
    return score


def set_field(
    row: dict[str, str],
    field: str,
    value: str,
    source_url: str,
    confidence: str,
    *,
    override: bool = False,
) -> None:
    value = clean(value)
    if not value:
        return
    if field == "website_url":
        value = root_url(value)
    if field in {"phone_number", "fax_number"}:
        value = normalize_phone(value)
        if not good_phone(value):
            return
    if field == "email_address" and not is_good_email(value):
        return
    if field == "fiscal_month":
        parsed = parse_fiscal_month(value)
        value = parsed or value

    current = clean(row.get(field, ""))
    if current and current != "対象外" and not override:
        return
    if current == "対象外" and not override:
        return
    row[field] = value
    source_column = SOURCE_COLUMNS.get(field)
    if source_column:
        row[source_column] = source_url
    confidence_column = CONFIDENCE_COLUMNS.get(field)
    if confidence_column:
        row[confidence_column] = confidence


def ensure_schema(row: dict[str, str]) -> dict[str, str]:
    for column in FINAL_FRONT_COLUMNS:
        row.setdefault(column, "")
    for field, source in SOURCE_COLUMNS.items():
        row.setdefault(field, "")
        row.setdefault(source, "")
    for confidence in CONFIDENCE_COLUMNS.values():
        row.setdefault(confidence, "")
    return row


def enrich_row(
    fetcher: Fetcher,
    row: dict[str, str],
    *,
    max_pages_per_site: int,
    max_pdfs_per_site: int,
    max_pdf_pages: int,
    search_web: bool,
) -> dict[str, str]:
    row = ensure_schema(dict(row))
    name = clean(row.get("company_name", ""))
    corporate_number = clean(row.get("corporate_number", ""))
    if not corporate_number:
        corporate_number = EXTRA_CORPORATE_NUMBERS.get(name, "")
        row["corporate_number"] = corporate_number

    entity_type = row.get("entity_type") or classify_entity(name)
    row["entity_type"] = entity_type
    for field in ["capital", "fiscal_month"]:
        reason = not_applicable(entity_type, field)
        if reason and not row.get(field):
            row[field] = reason

    nta = parse_nta_profile(fetcher, corporate_number)
    if nta.get("nta_source_url"):
        row["nta_source_url"] = nta.get("nta_source_url", "")
    row["nta_name"] = nta.get("nta_name", row.get("nta_name", ""))
    row["nta_last_updated"] = nta.get("nta_last_updated", row.get("nta_last_updated", ""))
    set_field(row, "detailed_address", nta.get("detailed_address", ""), nta.get("nta_source_url", ""), "official-db")

    gbiz = parse_gbiz_profile(fetcher, corporate_number)
    if gbiz.get("gbiz_source_url"):
        row["gbiz_source_url"] = gbiz.get("gbiz_source_url", "")
    set_field(row, "detailed_address", gbiz.get("本店所在地", ""), gbiz.get("gbiz_source_url", ""), "official-db")
    set_field(row, "website_url", gbiz.get("企業ホームページ", ""), gbiz.get("gbiz_source_url", ""), "official-db")
    set_field(row, "capital", gbiz.get("資本金", ""), gbiz.get("gbiz_source_url", ""), "official-db")
    set_field(row, "employee_count", gbiz.get("従業員数", ""), gbiz.get("gbiz_source_url", ""), "official-db")
    set_field(row, "fiscal_month", gbiz.get("決算月", ""), gbiz.get("gbiz_source_url", ""), "official-db")
    business = gbiz.get("事業概要", "")
    if gbiz.get("業種"):
        business = f"{gbiz['業種']} / {business}" if business else gbiz["業種"]
    set_field(row, "industry_business", business, gbiz.get("gbiz_source_url", ""), "official-db")

    irbank = parse_irbank(fetcher, corporate_number)
    row["irbank_mynumber_url"] = irbank.get("irbank_mynumber_url", row.get("irbank_mynumber_url", ""))
    row["irbank_company_url"] = irbank.get("irbank_company_url", row.get("irbank_company_url", ""))
    ir_source = irbank.get("irbank_company_url") or irbank.get("irbank_mynumber_url", "")
    listed = bool(irbank.get("irbank_company_url"))
    for field in ["website_url", "phone_number", "fiscal_month", "capital", "employee_count", "industry_business"]:
        set_field(row, field, irbank.get(field, ""), ir_source, "public-db", override=listed and field != "website_url")

    official_hint = known_official_url(name)
    if official_hint and not is_valid_http_url(row.get("website_url", "")):
        set_field(row, "website_url", official_hint, official_hint, "verified-hint", override=True)

    if not is_valid_http_url(row.get("website_url", "")) and search_web:
        url, source, title = search_web_for_official_site(fetcher, name)
        if url:
            set_field(row, "website_url", url, source, f"web-search:{title[:60]}", override=True)

    if row.get("website_url"):
        for source_url, data in crawl_official_site(fetcher, row["website_url"], max_pages_per_site):
            confidence = "official-site"
            for field, value in data.items():
                set_field(row, field, value, source_url, confidence)

        missing_pdf_fields = any(
            not row.get(field)
            for field in ["phone_number", "fax_number", "email_address", "capital", "employee_count", "fiscal_month", "industry_business"]
        )
        if missing_pdf_fields and max_pdfs_per_site > 0:
            for pdf_url in discover_pdf_urls(fetcher, row["website_url"], max_pages_per_site, max_pdfs_per_site):
                pdf_text = extract_pdf_text(fetcher, pdf_url, max_pdf_pages)
                if not pdf_text:
                    continue
                data = extract_fields_from_text(pdf_text)
                for field, value in data.items():
                    set_field(row, field, value, pdf_url, "official-pdf")

    finalize_row(row)
    return row


def finalize_row(row: dict[str, str]) -> None:
    if row.get("website_url"):
        row["website_url"] = root_url(row["website_url"])
    if row.get("phone_number") and not good_phone(row["phone_number"]):
        row["phone_number"] = ""
        row["phone_source_url"] = ""
        row["phone_confidence"] = ""
    if row.get("fax_number") and (not good_phone(row["fax_number"]) or row["fax_number"] == row.get("phone_number")):
        row["fax_number"] = ""
        row["fax_source_url"] = ""
        row["fax_confidence"] = ""
    if row.get("email_address") and not is_good_email(row["email_address"]):
        row["email_address"] = ""
        row["email_source_url"] = ""
        row["email_confidence"] = ""

    missing = [
        label
        for field, label in [
            ("phone_number", "電話"),
            ("fax_number", "FAX"),
            ("email_address", "メール"),
            ("contact_form_url", "問い合わせフォーム"),
            ("capital", "資本金"),
            ("employee_count", "従業員数"),
            ("fiscal_month", "決算月"),
            ("industry_business", "事業内容"),
        ]
        if not row.get(field)
    ]
    row["manual_check_reason"] = "公開Web自動調査では未確認: " + "、".join(missing) if missing else ""
    row["enrichment_status"] = "自動補完済み" if not missing else "一部未確認"


def load_input(path: Path, limit: int | None) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    if limit:
        rows = rows[:limit]
    return rows


def write_outputs(rows: list[dict[str, str]], output_prefix: Path, *, write_xlsx: bool) -> None:
    if not rows:
        return
    all_columns: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in all_columns:
                all_columns.append(key)
    columns = [c for c in FINAL_FRONT_COLUMNS if c in all_columns] + [c for c in all_columns if c not in FINAL_FRONT_COLUMNS]

    csv_path = output_prefix.with_suffix(".csv")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    if not write_xlsx:
        print(f"Wrote {csv_path}")
        return

    xlsx_path = output_prefix.with_suffix(".xlsx")
    df = pd.DataFrame(rows, columns=columns)
    df.to_excel(xlsx_path, index=False)
    format_xlsx(xlsx_path)
    print(f"Wrote {csv_path} and {xlsx_path}")


def format_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "enriched"
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col[:200])
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 70)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def print_stats(rows: list[dict[str, str]]) -> None:
    fields = [
        "detailed_address",
        "website_url",
        "phone_number",
        "fax_number",
        "email_address",
        "contact_form_url",
        "capital",
        "employee_count",
        "fiscal_month",
        "industry_business",
    ]
    total = len(rows)
    print("\nFill stats")
    for field in fields:
        filled = sum(1 for row in rows if row.get(field) and row.get(field) != "対象外")
        target_out = sum(1 for row in rows if row.get(field) == "対象外")
        suffix = f" (+対象外 {target_out})" if target_out else ""
        print(f"- {field}: {filled}/{total}{suffix}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape Soumu radio records and enrich all publicly crawlable company fields."
    )
    parser.add_argument(
        "--input",
        type=Path,
        help="Existing CSV to enrich. If omitted, the script scrapes the built-in Soumu q3 897.5MHz query.",
    )
    parser.add_argument("--limit", type=int, default=100, help="Maximum rows to process.")
    parser.add_argument(
        "--output-prefix",
        type=Path,
        default=Path("tele_soumu_oneclick_enriched"),
        help="Output path without extension.",
    )
    parser.add_argument("--max-pages-per-site", type=int, default=18, help="Maximum official-site pages per row.")
    parser.add_argument("--max-pdfs-per-site", type=int, default=2, help="Maximum official PDFs to parse per row.")
    parser.add_argument("--max-pdf-pages", type=int, default=8, help="Maximum pages to read from each PDF.")
    parser.add_argument("--sleep", type=float, default=0.35, help="Delay between HTTP requests.")
    parser.add_argument("--checkpoint-every", type=int, default=10, help="Write CSV checkpoint every N rows. 0 disables it.")
    parser.add_argument("--no-search-web", action="store_true", help="Do not use DuckDuckGo/Bing fallback for website discovery.")
    parser.add_argument("--no-xlsx", action="store_true", help="Only write CSV.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    fetcher = Fetcher(sleep_seconds=args.sleep)
    if args.input:
        rows = load_input(args.input, args.limit)
        print(f"Loaded {len(rows)} rows from {args.input}")
    else:
        rows = scrape_soumu_q3_897_5(fetcher, args.limit)
        print(f"Scraped {len(rows)} rows from Soumu")

    enriched: list[dict[str, str]] = []
    for index, row in enumerate(rows, 1):
        name = row.get("company_name", "")
        print(f"\n[{index}/{len(rows)}] {name}")
        enriched_row = enrich_row(
            fetcher,
            row,
            max_pages_per_site=args.max_pages_per_site,
            max_pdfs_per_site=args.max_pdfs_per_site,
            max_pdf_pages=args.max_pdf_pages,
            search_web=not args.no_search_web,
        )
        enriched.append(enriched_row)
        missing = enriched_row.get("manual_check_reason", "")
        print(f"done: website={enriched_row.get('website_url', '')} phone={enriched_row.get('phone_number', '')} {missing}")
        if args.checkpoint_every and index % args.checkpoint_every == 0:
            checkpoint = args.output_prefix.with_name(args.output_prefix.name + "_checkpoint")
            write_outputs(enriched, checkpoint, write_xlsx=False)

    write_outputs(enriched, args.output_prefix, write_xlsx=not args.no_xlsx)
    print_stats(enriched)


if __name__ == "__main__":
    main()
